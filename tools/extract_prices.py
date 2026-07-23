#!/usr/bin/env python3
"""
Build a price/analytics map from the GE price list (an .xlsx) and merge it into
the catalog.

The price list ("Прайс-лист ГЕ на согласование.xlsx") lists, per catalog
number (Артикул): an interchangeable article (Взаимозаменяемый артикул), a
Russian name, a price in CNY (без НДС) and a part group. These attributes — the
"аналитики" — are attached to catalog parts by matching the part number.

Outputs:
  catalog/data/prices.js            -> window.PRICES = { "<pn>": {p,g,x,n}, ... }
  catalog/data/all_part_numbers.csv -> unique catalog numbers WITH all analytics

Usage:
  pip install openpyxl
  python3 tools/extract_prices.py ["Прайс-лист ГЕ на согласование.xlsx"]
"""
import os
import re
import sys
import csv
import json
import glob

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_JS = os.path.join(ROOT, "catalog", "data", "parts.js")
PRICES_JS = os.path.join(ROOT, "catalog", "data", "prices.js")
CSV_OUT = os.path.join(ROOT, "catalog", "data", "all_part_numbers.csv")

# price-list column order (row 13): Артикул, Взаимозаменяемый артикул,
# Наименование, Цена CNY без НДС, Группа
COL_ART, COL_XREF, COL_NAME, COL_PRICE, COL_GROUP = 0, 1, 2, 3, 4


def norm_art(x):
    if x is None:
        return ""
    s = str(x).strip().replace("\xa0", " ").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def to_price(x):
    if x is None or x == "":
        return None
    try:
        return round(float(str(x).replace("\xa0", "").replace(",", ".")), 2)
    except ValueError:
        return None


def load_prices(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # find the header row (contains "Артикул")
    header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        if any(isinstance(c, str) and c.strip() == "Артикул" for c in row):
            header = i
            break
    if header is None:
        raise SystemExit("header row with 'Артикул' not found")

    prices = {}
    for row in ws.iter_rows(min_row=header + 1, values_only=True):
        art = norm_art(row[COL_ART])
        if not art:
            continue
        rec = {
            "p": to_price(row[COL_PRICE]),
            "g": (str(row[COL_GROUP]).strip() if row[COL_GROUP] not in (None, "") else ""),
            "x": norm_art(row[COL_XREF]),
            "n": (str(row[COL_NAME]).replace("\xa0", " ").strip() if row[COL_NAME] not in (None, "") else ""),
        }
        prices.setdefault(art, rec)
        # also index by the interchangeable article so a catalog number listed
        # only as a cross-reference still resolves
        if rec["x"]:
            prices.setdefault(rec["x"], dict(rec))
    return prices


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else None
    if not src:
        cand = glob.glob(os.path.join(ROOT, "*райс*.xlsx")) + glob.glob(os.path.join(ROOT, "*.xlsx"))
        if not cand:
            raise SystemExit("price-list .xlsx not found; pass its path")
        src = cand[0]

    prices = load_prices(src)

    data = json.loads(open(DATA_JS, encoding="utf-8").read()[len("window.CATALOG = "):-2])
    catalog_pns = set()
    for s in data["sections"]:
        for f in s["figures"]:
            for p in f["parts"]:
                if p["pn"]:
                    catalog_pns.add(p["pn"])

    # keep only price records referenced by the catalog (smaller payload)
    used = {pn: prices[pn] for pn in catalog_pns if pn in prices}
    matched = sum(1 for r in used.values() if r["p"] is not None)

    with open(PRICES_JS, "w", encoding="utf-8") as fh:
        fh.write("window.PRICES = ")
        json.dump(used, fh, ensure_ascii=False, separators=(",", ":"))
        fh.write(";\n")

    # ---- unique catalog numbers WITH all analytics ----
    uniq = {}
    for s in data["sections"]:
        src_tag = "GE" if s["chapter"] == "600" else "NTE200"
        for f in s["figures"]:
            for p in f["parts"]:
                if not p["pn"]:
                    continue
                u = uniq.setdefault(p["pn"], {"pn": p["pn"], "zh": p["zh"], "en": p["en"],
                                              "secs": set(), "src": set()})
                u["secs"].add(s["code"])
                u["src"].add(src_tag)
                if not u["en"] and p["en"]:
                    u["en"] = p["en"]
                if not u["zh"] and p["zh"]:
                    u["zh"] = p["zh"]

    with open(CSV_OUT, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Артикул (Part No.)", "Наименование (RU)", "Description (EN)",
                    "Description (ZH)", "Цена, CNY без НДС", "Группа",
                    "Взаимозаменяемый артикул", "Источник", "Разделы"])
        for pn in sorted(uniq):
            u = uniq[pn]
            pr = prices.get(pn, {})
            w.writerow([pn, pr.get("n", ""), u["en"], u["zh"],
                        ("" if pr.get("p") is None else pr.get("p")),
                        pr.get("g", ""), pr.get("x", ""),
                        "/".join(sorted(u["src"])), " ".join(sorted(u["secs"]))])

    print("Price rows loaded: %d" % len(prices))
    print("Catalog part numbers: %d  with a price: %d" % (len(catalog_pns), matched))
    print("Wrote %s (%d records)" % (os.path.relpath(PRICES_JS, ROOT), len(used)))
    print("Wrote %s (%d unique numbers)" % (os.path.relpath(CSV_OUT, ROOT), len(uniq)))


if __name__ == "__main__":
    main()
